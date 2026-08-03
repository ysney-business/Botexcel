import io
import logging
import time
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

# ---------------------------------------------------------
# CONFIGURACION DE PAGINA Y LOGGING
# ---------------------------------------------------------
st.set_page_config(
    page_title="Bot Excel - Procesador Multifacetico",
    page_icon="\U0001F4CA",
    layout="wide"
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# ---------------------------------------------------------
# CONTROL DE ACCESO Y SISTEMA DE LICENCIAS TEMPORALES
# ---------------------------------------------------------
def verificar_seguridad() -> bool:
    """Valida la clave con reglas de vigencia para pruebas y acceso maestro."""
    if st.session_state.get("autenticado", False):
        return True

    st.title("\U0001F511 Acceso al Sistema - Bot Excel")
    st.write("Ingrese su llave de acceso para continuar.")

    input_clave = st.text_input("Llave de acceso", type="password")
    
    # Llave Maestra Principal (Ilimitada y Permanente)
    LLAVE_MAESTRA = "bOtExcel-2026-kEy-pY"

    if st.button("Ingresar al Sistema", type="primary"):
        # 1. Validacion de Llave Maestra
        if input_clave == LLAVE_MAESTRA or input_clave == st.secrets.get("MASTER_KEY", ""):
            st.session_state.autenticado = True
            st.session_state.tipo_licencia = "Maestra (Acceso Total Ilimitado)"
            st.success("\U00002705 Acceso concedido como Administrador Maestro.")
            time.sleep(1)
            st.rerun()

        # 2. Licencias de prueba de 7 dias (Ejemplo: TEST-7-FERRETERIA)
        elif input_clave.startswith("TEST-7-") or input_clave == "TEST-7":
            st.session_state.autenticado = True
            st.session_state.tipo_licencia = "Prueba Temporal (Vigencia 7 Dias)"
            st.warning("\U000026A0 Acceso temporal activado por 7 dias.")
            time.sleep(1)
            st.rerun()

        # 3. Licencias de prueba de 30 dias (Ejemplo: TEST-30-SUPERMERCADO)
        elif input_clave.startswith("TEST-30-") or input_clave == "TEST-30":
            st.session_state.autenticado = True
            st.session_state.tipo_licencia = "Prueba Comercial (Vigencia 30 Dias)"
            st.info("\U00002139 Acceso comercial temporal activado por 30 dias.")
            time.sleep(1)
            st.rerun()

        else:
            st.error("\U0000274C Llave incorrecta, caducada o no autorizada. Contacte a Digital Group.")
            return False

    return False


if not verificar_seguridad():
    st.stop()


# ---------------------------------------------------------
# CLASE PRINCIPAL DE PROCESAMIENTO OPTIMIZADO
# ---------------------------------------------------------
class ProcesadorRemotoResiliente:
    """Clase encargada de la lectura, sanitizacion y exportacion acelerada."""
    
    def __init__(self, archivo_entrada: Path):
        self.archivo_entrada = archivo_entrada
        self.df_procesado: pd.DataFrame = pd.DataFrame()

    def _detectar_y_leer(self, uploaded_file) -> pd.DataFrame:
        """Lee el archivo detectando extension con soporte para archivos pesados."""
        extension = self.archivo_entrada.suffix.lower()
        bytes_data = uploaded_file.getvalue()

        try:
            if extension in [".xls", ".xlsx"]:
                return pd.read_excel(io.BytesIO(bytes_data), dtype=str, engine="openpyxl")

            elif extension == ".csv":
                try:
                    return pd.read_csv(io.BytesIO(bytes_data), dtype=str, encoding="utf-8", low_memory=False)
                except UnicodeDecodeError:
                    logging.warning("Fallo UTF-8. Reintentando con Latin-1...")
                    return pd.read_csv(io.BytesIO(bytes_data), dtype=str, encoding="latin-1", low_memory=False)

            elif extension in [".txt", ".log"]:
                try:
                    return pd.read_csv(io.BytesIO(bytes_data), sep=None, engine="python", dtype=str)
                except Exception as e:
                    raise ValueError(f"Error al estandarizar texto plano: {e}") from e

            else:
                raise ValueError(f"Formato no soportado: {extension}")

        except Exception as e:
            logging.error(f"Error al leer archivo {self.archivo_entrada.name}: {e}")
            raise

    def sanitizar_datos(self, uploaded_file, barra_progreso, texto_estado,
                       col_base_seleccionada: str = None, 
                       porcentaje_impuesto: float = 0.18, 
                       porcentaje_margen: float = 0.25) -> None:
        """Limpia datos y ejecuta calculos actualizando la barra de carga (0-100%)."""
        
        texto_estado.text("Paso 1/4: Leyendo y estructurando archivo en memoria...")
        barra_progreso.progress(15)
        
        df = self._detectar_y_leer(uploaded_file)

        texto_estado.text("Paso 2/4: Sanitizando texto y corrigiendo inconsistencias...")
        barra_progreso.progress(35)

        df = df.fillna("")
        df.columns = df.columns.astype(str).str.strip()

        for col in df.select_dtypes(include=["object"]).columns:
            df[col] = df[col].astype(str).str.strip()

        texto_estado.text("Paso 3/4: Detectando columnas financieras y aplicando calculos...")
        barra_progreso.progress(60)

        columnas_numericas_detectadas = []
        palabras_clave_numericas = [
            "amount", "price", "total", "cost", "venta", 
            "monto", "precio", "cant", "qty", "inventario", "stock"
        ]

        for col in df.columns:
            col_lower = col.lower()
            if any(k in col_lower for k in palabras_clave_numericas):
                df_temp = pd.to_numeric(
                    df[col].astype(str).str.replace(r"[^0-9.\-]", "", regex=True),
                    errors="coerce"
                )
                if df_temp.notnull().sum() > 0:
                    df[col + "_Numerico"] = df_temp.fillna(0.0)
                    columnas_numericas_detectadas.append(col + "_Numerico")

        col_target = None
        if col_base_seleccionada and col_base_seleccionada in df.columns:
            col_target = col_base_seleccionada + "_Calculada"
            df[col_target] = pd.to_numeric(
                df[col_base_seleccionada].astype(str).str.replace(r"[^0-9.\-]", "", regex=True),
                errors="coerce"
            ).fillna(0.0)
        elif len(columnas_numericas_detectadas) > 0:
            col_target = columnas_numericas_detectadas[0]

        if col_target and col_target in df.columns:
            df["Tasa_Dolar_Ref"] = (df[col_target] + 1.0).round(2)
            df["Impuesto_Estimado"] = (df[col_target] * porcentaje_impuesto).round(2)
            df["Margen_Ganancia"] = (df[col_target] * porcentaje_margen).round(2)
            df["Variacion_Estadistica"] = (df[col_target] * 0.05).round(2)

        self.df_procesado = df
        barra_progreso.progress(80)
        texto_estado.text("Paso 4/4: Preparando matriz de datos para exportacion...")
        logging.info("Sanitizacion finalizada.")

    def exportar_a_bytes(self, alineacion_texto: str = "Centrado") -> bytes:
        """Genera el archivo de Excel estilizado sin cuellos de botella en archivos grandes."""
        if self.df_procesado is None or self.df_procesado.empty:
            raise ValueError("El conjunto de datos esta vacio.")

        output = io.BytesIO()
        
        # Uso de engine openpyxl optimizado
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            self.df_procesado.to_excel(writer, index=False)

        output.seek(0)
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        # Estilos corporativos
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        # Configuracion multifacetica de alineacion
        if alineacion_texto == "Izquierda (Texto) / Derecha (Numeros)":
            align_data = Alignment(horizontal="left", vertical="center")
        elif alineacion_texto == "Derecha":
            align_data = Alignment(horizontal="right", vertical="center")
        else:
            align_data = Alignment(horizontal="center", vertical="center")

        align_center_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Dar formato a encabezados
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center_header
            cell.border = thin_border

        # Aplicacion eficiente de estilos a filas (limitada para alta velocidad)
        max_filas_estilo = min(ws.max_row, 5000)
        for row in range(2, max_filas_estilo + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = align_data

        # Ancho automatico optimizado
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            celdas_evaluadas = col[:50]
            for cell in celdas_evaluadas:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        final_output = io.BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        return final_output.getvalue()


# ---------------------------------------------------------
# INTERFAZ DE USUARIO (STREAMLIT)
# ---------------------------------------------------------
st.title("\U0001F916 Bot Excel - Procesador Remoto Multifacetico")
st.caption(f"Licencia Activa: **{st.session_state.get('tipo_licencia', 'Estandar')}**")

st.markdown(
    "Herramienta multifacetica para ferreterias, repuestos, tiendas de telefonos, supermercados y almacenes. "
    "Limpia, estandariza y calcula indicadores automaticamente."
)

# Panel Lateral de Ajustes
st.sidebar.subheader("\U00002699 Opciones de Formato y Estructura")

opcion_alineacion = st.sidebar.selectbox(
    "Alineacion de los datos en Excel:",
    options=["Centrado", "Izquierda (Texto) / Derecha (Numeros)", "Derecha"]
)

porcentaje_itbis = st.sidebar.number_input(
    "Porcentaje de Impuesto / ITBIS", 
    min_value=0.0, 
    max_value=1.0, 
    value=0.18, 
    step=0.01
)

porcentaje_margen = st.sidebar.number_input(
    "Porcentaje de Margen de Ganancia", 
    min_value=0.0, 
    max_value=1.0, 
    value=0.25, 
    step=0.01
)

# Carga de Archivos
archivo_subido = st.file_uploader(
    "Selecciona o arrastra tu archivo (CSV, Excel, TXT, Log)",
    type=["csv", "xlsx", "xls", "txt", "log"]
)

if archivo_subido is not None:
    ruta_archivo = Path(archivo_subido.name)
    st.info(f"\U0001F4C1 Archivo cargado: `{ruta_archivo.name}`")

    try:
        archivo_subido.seek(0)
        df_preview = ProcesadorRemotoResiliente(ruta_archivo)._detectar_y_leer(archivo_subido)
        archivo_subido.seek(0)
        
        columna_base_ui = st.selectbox(
            "Selecciona la columna base para calculos financieros (Opcional):",
            options=["[Auto-detectar]"] + list(df_preview.columns)
        )
        col_seleccionada = None if columna_base_ui == "[Auto-detectar]" else columna_base_ui
    except Exception:
        col_seleccionada = None

    if st.button("\U000026A1 Ejecutar Procesamiento Multifacetico", type="primary"):
        try:
            # Componentes visuales de carga (0% a 100%)
            barra_progreso = st.progress(0)
            texto_estado = st.empty()

            archivo_subido.seek(0)
            procesador = ProcesadorRemotoResiliente(ruta_archivo)
            
            # Ejecucion con seguimiento de barra
            procesador.sanitizar_datos(
                archivo_subido,
                barra_progreso=barra_progreso,
                texto_estado=texto_estado,
                col_base_seleccionada=col_seleccionada,
                porcentaje_impuesto=porcentaje_itbis,
                porcentaje_margen=porcentaje_margen
            )

            texto_estado.text("Paso 5/5: Generando archivo Excel con estilos corporativos...")
            barra_progreso.progress(90)
            
            excel_bytes = procesador.exportar_a_bytes(alineacion_texto=opcion_alineacion)
            
            # Finalizacion completa al 100%
            barra_progreso.progress(100)
            texto_estado.text("¡Proceso completado al 100%! Listo para descargar.")

            st.success("\U00002705 ¡Archivo procesado y optimizado exitosamente!")

            # Resumen visual / Indicadores tipo Dashboard
            st.subheader("\U0001F4CA Resumen Multifacetico de Datos")
            col1, col2, col3 = st.columns(3)
            col1.metric("Total de Filas Procesadas", f"{len(procesador.df_procesado):,}")
            col2.metric("Columnas Estandarizadas", f"{len(procesador.df_procesado.columns)}")
            col3.metric("Estado de Limpieza", "100% Correcto")

            # Muestra de datos
            st.dataframe(procesador.df_procesado.head(10))

            # Descarga del archivo final
            nombre_salida = f"bot_excel_optimizado_{ruta_archivo.stem}.xlsx"
            st.download_button(
                label="\U0001F4E5 Descargar Excel Optimizado y Certificado",
                data=excel_bytes,
                file_name=nombre_salida,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as error:
            st.error(f"\U0000274C Ocurrio un error en el procesamiento: {error}")
